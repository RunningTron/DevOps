from openpyxl import load_workbook,Workbook

wb = Workbook()
ws = wb.active
wb.create_sheet()


'''
# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook('pythontab.xlsx')

# 默认打开的文件为可读写，若有需要可以指定参数read_only为True。

# 获取工作表--Sheet
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet1')
# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active

# 获取单格
# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['B4']
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}') # 返回的数字就是int型
# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B2
b4_too = sheet.cell(row=4, column=2)
print(b4_too.value)
'''
class A:
    c= 1
    def b(self):
        pass

a = A()
id(a.c)